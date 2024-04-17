import glob
import os
import re
import string
from datetime import datetime, timedelta
import openpyxl
from jira import JIRA

JIRA_PROJCET = "SAW"  # "SAW"  "TEST"
# JIRA_PROJCET = "TEST"


JIRA_URL = "https://jiratepz1.mobiledrivetech.com/"
USERNAME = "qaautotest"
PASSWORD = "MobileDrive#01"
AUTH_INFO = (USERNAME, PASSWORD)

SEARCH_PATTERN = "project = SAW AND reporter = qaautotest"

SAW_DIR = r"\\MD-TEST\Projects\SAW"
LOCAL_DIR = r"D:\Projects\SAW"
# \\MD-TEST\Projects\SAW\Monkey_log\2023_08_01\outputFolder\2023_08_01-03_30_03
TODAY = datetime.now()
TODAY_DATE = TODAY.strftime("%Y_%m_%d")


def delete_exist_excel(dir_path):
    for node in os.listdir(dir_path):
        print(node)
        if os.path.isfile(dir_path + os.sep + node) and node.startswith("Monkey_issues_"):
            print(node)
            os.remove(dir_path + os.sep + node)


def find_issue_in_jira():
    jira = JIRA(JIRA_URL, auth=AUTH_INFO)
    search_result = jira.search_issues(SEARCH_PATTERN, maxResults=False)
    return search_result


def write_excel_file(dir_path, issue_list):
    try:
        delete_exist_excel(dir_path)
    except PermissionError:
        print("Failed to delete old file(s)")
    workbook = openpyxl.Workbook()
    issue_sheet = workbook.worksheets[0]
    issue_sheet.title = "issue_list"
    title_list = ["Issue key", "Issue id", "Status", "Created", "Updated", "Summary", "Description"]
    letters = list(string.ascii_lowercase)

    for column in range(len(title_list)):
        issue_sheet[f"{letters[column]}1"].value = title_list[column]

    for row in range(len(issue_list)):
        issue_sheet.insert_rows(2)
        issue = issue_list[row]
        tz_regular = timedelta(hours=8)
        created = datetime.strftime(datetime.strptime(issue.fields.created, '%Y-%m-%dT%H:%M:%S.000+0000')+tz_regular,
                                    "%Y-%m-%d %H:%M:%S")
        updated = datetime.strftime(datetime.strptime(issue.fields.updated, '%Y-%m-%dT%H:%M:%S.000+0000'),
                                    "%Y-%m-%d %H:%M:%S")
        issue_fields = [issue.key, issue.id, issue.fields.status.name,
                        created, updated, issue.fields.summary, issue.fields.description]
        for column in range(len(title_list)):
            insert_cell = issue_sheet[f"{letters[column]}2"]
            insert_cell.value = issue_fields[column]

    workbook.save(dir_path + os.sep + f"Monkey_issues_{TODAY_DATE}.xlsx")


def parsing_issue_counts(dir_path):
    workbook = openpyxl.load_workbook(dir_path + os.sep + f"Monkey_issues_{TODAY_DATE}.xlsx")
    issue_log = workbook.create_sheet("issue_frequency_count", 1)
    monkey_log_folder = os.path.abspath(dir_path + os.sep + "Monkey_log")
    monkey_log_list = os.listdir(monkey_log_folder)
    log_list = {}
    for folder in monkey_log_list:
        if re.match("[0-9_]+_[a-z.]+", folder):
            # print(folder)
            text_file_list = glob.glob(
                monkey_log_folder + os.sep + folder + os.sep + "outputFolder" + os.sep + "*")
            for output_folder in text_file_list:
                if all(os.path.exists(output_folder + os.sep + file) for file in
                       ["Build_info.txt", "JIRA_result_output.txt"]):
                    print(folder)
                    build_data = open(output_folder + os.sep + "Build_info.txt", "r").read()
                    build_info = re.search(r"(.+)", build_data).group(1)
                    print(build_info)
                    jira_data = open(output_folder + os.sep + "JIRA_result_output.txt", "r").read()
                    if build_info not in log_list.keys():
                        log_list[f"{build_info}"] = {}
                    for issue, count in re.findall("([A-Z]+-[0-9]+) x([0-9]+)", jira_data):
                        # print(issue, count)
                        if issue not in log_list[f"{build_info}"].keys():
                            log_list[f"{build_info}"][f"{issue}"] = int(count)
                        else:
                            log_list[f"{build_info}"][f"{issue}"] += int(count)
    print(log_list)
    title_list = ["Issue key"]
    letters = list(string.ascii_lowercase)
    for column in range(len(title_list)):
        issue_log[f"{letters[column]}1"].value = title_list[column]
    issue_list = workbook["issue_list"]
    for row in range(1, issue_list.max_row + 1):
        issue_log[f"A{row}"].value = issue_list[f"A{row}"].value

    for row in log_list.keys():
        issue_log.insert_cols(2)
        issue_log["B1"].value = row
        for i in range(issue_log.max_row):
            for data in log_list[row].keys():
                if issue_log[f"A{i + 2}"].value == data:
                    issue_log[f"B{i + 2}"].value = log_list[row][data]
    workbook.save(dir_path + os.sep + f"Monkey_issues_{TODAY_DATE}.xlsx")


if __name__ == '__main__':
    if os.path.exists(LOCAL_DIR + os.sep + "SAW_Auto_Weekly Report.xlsx"):
        file_folder = LOCAL_DIR
    else:
        file_folder = SAW_DIR
    write_excel_file(file_folder, find_issue_in_jira())
    # parsing_issue_counts(SAW_DIR)
