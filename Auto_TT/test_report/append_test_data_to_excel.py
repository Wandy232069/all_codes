import datetime
import glob
import os
import re
import string
import openpyxl
import openpyxl.utils.datetime
import json

SAW_DIR = r"\\MD-TEST\Projects\SAW"
LOCAL_DIR = r"D:\Projects\SAW"
# \\MD-TEST\Projects\SAW\Monkey_log\2023_08_01\outputFolder\2023_08_01-03_30_03
TODAY = datetime.datetime.now()
TODAY_DATE = TODAY.strftime("%Y_%m_%d")


def write_weekly_report_monkey(dir_path, log_date):
    dir_path = os.path.abspath(dir_path)
    formatted_log_date = log_date.strftime("%Y_%m_%d")
    print(formatted_log_date)

    try:
        data_dir_list = \
            glob.glob(dir_path + os.sep + "Monkey_log" + os.sep + f"{formatted_log_date}*" + os.sep + "outputFolder")
        print(data_dir_list)
    except IndexError:
        print("no output dir found:" + formatted_log_date)
        return
    for data_dir in data_dir_list:
        if os.path.isdir(data_dir):
            for log_dir in os.listdir(data_dir):
                print(data_dir + os.sep + log_dir)
                if os.path.exists(data_dir + os.sep + log_dir + os.sep + "JIRA_result_output.txt"):
                    # Count the number of New issues and Duplicate issues
                    with open(data_dir + os.sep + log_dir + os.sep + "JIRA_result_output.txt", "r") as jira_file:
                        jira_data = jira_file.read()
                        new_issues_count = re.findall(r"New issues x ([0-9]+):\n", jira_data)[0]
                        duplicate_issues_count = re.findall(r"Duplicate issues x ([0-9]+):\n", jira_data)[0]
                if all(os.path.exists(data_dir + os.sep + log_dir + os.sep + file) for file in ["Build_info.txt",
                                                                                                "Test_packages.txt",
                                                                                                "CrashReport.txt",
                                                                                                "Execution_time.txt"]):
                    with open(data_dir + os.sep + log_dir + os.sep + "Build_info.txt", "r") as build_info_file:
                        build_info = build_info_file.read()

                    with open(data_dir + os.sep + log_dir + os.sep + "Test_packages.txt", "r") as test_apps_file:
                        test_apps = test_apps_file.read()

                    with open(data_dir + os.sep + log_dir + os.sep + "Execution_time.txt", "r") as execution_time_file:
                        execution_time_data = execution_time_file.read()
                        execution_time_str = re.findall(r"Execution duration : (.+)", execution_time_data)
                        if execution_time_str:
                            matches = re.findall(r'(\d+)H:(\d+)M:(\d+)S', execution_time_str[0])
                            if matches:
                                hours, minutes, seconds = map(int, matches[0])
                                total_hours = hours + minutes / 60 + seconds / 3600
                            else:
                                total_hours = 0
                        else:
                            total_hours = None
                        execution_time = total_hours

                    with open(data_dir + os.sep + log_dir + os.sep + "CrashReport.txt", "r") as crash_data_file:
                        crash_data = crash_data_file.read()
                        anr_count = re.findall(r"There are ([0-9]+) ANR crashes in logs", crash_data)[0]
                        try:
                            md_anr_count = re.findall(r"There are ([0-9]+) MD's ANR crashes in logs", crash_data)[0]
                        except IndexError:
                            md_anr_count = 0
                        dropbox_app_crash_count = \
                            re.findall(r"There are ([0-9]+) system_app crashes in logs", crash_data)[0]
                        dropbox_system_crash_count = \
                            re.findall(r"There are ([0-9]+) system_server crashes in logs", crash_data)[0]
                        try:
                            md_dropbox_count = \
                                re.findall(r"There are ([0-9]+) MD's app crashes in logs", crash_data)[0]
                        except IndexError:
                            md_dropbox_count = 0
                        try:
                            md_failed_count = int(md_anr_count) + int(md_dropbox_count)
                            if execution_time:
                                mtbf = execution_time / (md_failed_count + 1)
                            else:
                                mtbf = None
                        except UnboundLocalError:
                            md_failed_count = None
                            mtbf = None
                else:
                    break
                workbook = openpyxl.load_workbook(dir_path + os.sep + "SAW_Auto_Weekly Report.xlsx")
                monkey_sheet = workbook['MonkeyAllHistory']
                letters = list(string.ascii_lowercase)

                exist_date_list = []
                for rows in monkey_sheet.iter_cols(max_col=1, min_row=2, max_row=monkey_sheet.max_row):
                    for cell in rows:
                        try:
                            exist_date_list.append(cell.value.strftime("%Y_%m_%d-%H_%M_%S"))
                        except Exception:
                            exist_date_list.append(cell.value)

                print(exist_date_list)

                if log_dir not in exist_date_list:
                    flag = 0
                else:
                    if monkey_sheet[f"C{exist_date_list.index(str(log_dir)) + 2}"].value != test_apps and \
                            monkey_sheet[
                                f"A{exist_date_list.index(str(log_dir)) + 2}"].value != \
                            datetime.datetime.strptime(log_dir, "%Y_%m_%d-%H_%M_%S"):
                        print(monkey_sheet[f"C{exist_date_list.index(str(log_dir)) + 2}"].value)
                        print(test_apps)
                        print(monkey_sheet[
                                  f"A{exist_date_list.index(str(log_dir)) + 2}"].value)
                        print(datetime.datetime.strptime(log_dir, "%Y_%m_%d-%H_%M_%S"))
                        flag = 0
                    else:
                        flag = 1

                if flag == 0:
                    monkey_sheet.insert_rows(2)
                    for column_count in range(monkey_sheet.max_column):
                        title = monkey_sheet[f"{letters[column_count]}1"].value
                        insert_cell = monkey_sheet[f"{letters[column_count]}2"]
                        if title == "Date":
                            insert_cell.value = datetime.datetime.strptime(log_dir, "%Y_%m_%d-%H_%M_%S")
                            insert_cell.number_format = "m/d;@"
                        elif title == "Execution Time":
                            insert_cell.value = execution_time
                            insert_cell.number_format = "0.00"
                        elif title == "New Issues":
                            insert_cell.value = int(new_issues_count)
                            insert_cell.number_format = '0'
                        elif title == "Known Issues":
                            insert_cell.value = int(duplicate_issues_count)
                            insert_cell.number_format = "0"
                        elif title == "MD_ANR":
                            insert_cell.value = int(md_anr_count)
                            insert_cell.number_format = "0"
                        elif title == "ANR":
                            insert_cell.value = int(anr_count)
                            insert_cell.number_format = "0"
                        elif title == "MD_App Crash":
                            insert_cell.value = int(md_dropbox_count)
                            insert_cell.number_format = "0"
                        elif title == "System_App Crash":
                            insert_cell.value = int(dropbox_app_crash_count)
                            insert_cell.number_format = "0"
                        elif title == "System_Server Crash":
                            insert_cell.value = int(dropbox_system_crash_count)
                            insert_cell.number_format = "0"
                        elif title == "APK & Version":
                            insert_cell.value = test_apps
                        elif title == "Build Version":
                            insert_cell.value = build_info
                        elif title == "Failed Count":
                            insert_cell.value = md_failed_count
                            insert_cell.number_format = "0"
                        elif title == "MTBF":
                            insert_cell.value = mtbf
                            insert_cell.number_format = "0.00"
                        elif title == "DUTs":
                            insert_cell.value = 1
                            insert_cell.number_format = "0"

                else:
                    # print(formatted_log_date + "Date data exist")
                    print("duplicate data in row" + str(exist_date_list.index(str(log_dir)) + 2))
                try:
                    workbook.save(dir_path + os.sep + "SAW_Auto_Weekly Report.xlsx")
                except PermissionError:
                    print("Workbook is in use")
                    workbook.save(dir_path + os.sep + f"SAW_Auto_Weekly Report_{TODAY_DATE}.xlsx")


def write_weekly_report_pytest(dir_path, log_date):
    dir_path = os.path.abspath(dir_path) + os.sep + "Pytest_log"
    formatted_log_date = log_date.strftime("%Y%m%d")
    print(formatted_log_date)

    for file in os.listdir(dir_path):
        if formatted_log_date in file and file.endswith(".json"):
            json_file = open(dir_path + os.sep + file, 'r')
            data = json.loads(json_file.read())
            json_file.close()
            total_tests = data["total_tests"]
            try:
                total_time = data["total_time"]
            except KeyError:
                total_time = "00:00:00"
            pass_count = data["status_list"]['pass']
            fail_count = data["status_list"]['fail']
            print(f"total:{total_tests}, pass:{pass_count}, fail:{fail_count}, total_time:{total_time}")
            # write data to workbook
            workbook_path = os.path.dirname(dir_path) + os.sep + "SAW_Auto_Weekly Report.xlsx"
            workbook = openpyxl.load_workbook(workbook_path)
            monkey_sheet = workbook['StressAllHistory']
            letters = list(string.ascii_lowercase)

            exist_date_list = []
            for rows in monkey_sheet.iter_cols(max_col=1, min_row=2, max_row=monkey_sheet.max_row):
                for cell in rows:
                    try:
                        exist_date_list.append(cell.value.strftime("%Y%m%d"))
                    except Exception:
                        exist_date_list.append(cell.value)

            print(exist_date_list)
            if str(formatted_log_date) not in exist_date_list:
                monkey_sheet.insert_rows(2)
                for column in range(monkey_sheet.max_column):
                    title = monkey_sheet[f"{letters[column]}1"].value
                    insert_cell = monkey_sheet[f"{letters[column]}2"]
                    if title == "Date":
                        insert_cell.value = log_date
                        insert_cell.number_format = "m/d;@"
                    elif title == "Total Test":
                        insert_cell.value = total_tests
                    elif title == "Pass":
                        insert_cell.value = pass_count
                    elif title == "Fail":
                        insert_cell.value = fail_count
                    elif title == "Running time (hrs.)":
                        total_time_hr = re.findall("([0-9]+)", total_time)[0]
                        total_time_min = re.findall("([0-9]+)", total_time)[1]
                        total_time_hr = int(total_time_hr) + round(int(total_time_min) / 60, 3)
                        insert_cell.value = total_time_hr

            else:
                print(f"{formatted_log_date} exist\n")
            saw_dir = os.path.dirname(dir_path)
            try:
                workbook.save(saw_dir + os.sep + "SAW_Auto_Weekly Report.xlsx")
            except PermissionError:
                print("Workbook is in use")
                workbook.save(saw_dir + os.sep + f"SAW_Auto_Weekly Report_{TODAY_DATE}.xlsx")


if __name__ == "__main__":
    if os.path.exists(LOCAL_DIR + os.sep + "SAW_Auto_Weekly Report.xlsx"):
        file_path = LOCAL_DIR
    else:
        file_path = SAW_DIR
    for i in range(7):
        date = TODAY - datetime.timedelta(days=6 - i)
        print("Monkey")
        write_weekly_report_monkey(file_path, date)
        print("Pytest")
        write_weekly_report_pytest(file_path, date)
