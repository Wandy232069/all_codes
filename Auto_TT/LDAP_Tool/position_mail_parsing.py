import json
import os
import re
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import ldap_authenticator
import openpyxl
from DB_tool.db_tool import myDB

data = []

FILE_PATH = r"/Users/larryhzlai/Downloads/24PI2 Scrum list_delete.xlsx"
ldap = ldap_authenticator.LDAPAuthenticator()
workbook = openpyxl.load_workbook(FILE_PATH)
sheet = workbook.worksheets[0]
# r"(cn=LarryHZLai\28賴昊澤\29)"
for j in range(1, sheet.max_column + 1):
    position = sheet.cell(row=1, column=j + 1).value
    print(sheet.cell(row=1, column=j).value)
    for i in range(2, sheet.max_row + 1):
        module = sheet.cell(row=i, column=1).value
        print(sheet.cell(row=i, column=1).value)
        cell_value = sheet.cell(row=i, column=j + 1).value
        name_list = str(cell_value).split("\n")
        for name in name_list:
            if name == "" or name == "None":
                break
            for keyword in ["(TP)", "(主)", "(IN)", "(Half)", "(AA)"]:
                name = name.replace(keyword, "")
            print(name)
            cn_chinese = re.sub('[a-zA-Z ()]', "", name)
            print(cn_chinese)
            if len(cn_chinese) > 1:
                cn_ = "(displayName=*" + cn_chinese[0:2] + "*)"

            else:
                cn_ = "(displayName=" + re.match(r"([a-zA-Z. ]+)", name).group(1) + "*)"
                # result_list = [{"mail": ".".join(re.findall(r"([a-zA-Z]+)", name))+"@mobiledrivetech.com"}]
            print(cn_)
            result_list = ldap.get_subtree_entries(search_filter=cn_,
                                                   search_base="OU=Users-mobiledrivetech,"
                                                               "OU=R30007,DC=fihtdc,"
                                                               "DC=com",
                                                   attributes=["mail"])
            for result in result_list:
                if all(match in name for match in re.findall(r"([a-zA-Z]+)", name)[0:1]):
                    print(result["mail"])
                    mail = str(result["mail"])
                    data.append({"module": module, "position": position, "name": name, "mail": mail})
print(data)
with open("result.json", "w") as json_file:
    json_file.write(json.dumps(data, separators=(',', ':')))

db = myDB("mail_group")
for item in data:
    db.insert_data("saw_mail_group", ["platform", "module", "position", "name", "mail"],
                   ["AOSP", item["module"], item["position"], item["name"], item["mail"]])
